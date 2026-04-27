import csv
import io
from datetime import date, datetime
from typing import Any

from raggae.domain.exceptions.document_exceptions import DocumentExtractionError


class TabularDocumentTextExtractor:
    """Extract text from CSV, XLSX, and XLS files into structured tabular format."""

    async def extract_text(self, file_name: str, content: bytes, content_type: str) -> str:
        _ = content_type
        extension = file_name.rsplit(".", maxsplit=1)[-1].lower() if "." in file_name else ""

        if extension == "csv":
            return self._extract_csv(content)
        if extension == "xlsx":
            return self._extract_xlsx(content)
        if extension == "xls":
            return self._extract_xls(content)
        raise DocumentExtractionError(f"Unsupported tabular extension: {extension}")

    def _extract_csv(self, content: bytes) -> str:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        sample = text[:4096]
        try:
            dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel

        rows: list[list[str]] = []
        for row in csv.reader(io.StringIO(text), dialect):
            rows.append([cell.strip() for cell in row])
        return self._rows_to_structured_text(rows, sheet_name=None)

    def _extract_xlsx(self, content: bytes) -> str:
        try:
            import openpyxl
        except ModuleNotFoundError as exc:
            raise DocumentExtractionError("openpyxl is required for XLSX extraction") from exc

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            parts: list[str] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = [[self._format_cell_value(cell.value) for cell in row] for row in ws.iter_rows()]
                sheet_text = self._rows_to_structured_text(rows, sheet_name=sheet_name)
                if sheet_text:
                    parts.append(sheet_text)
            return "\n".join(parts)
        except DocumentExtractionError:
            raise
        except Exception as exc:
            raise DocumentExtractionError(f"Failed to extract XLSX: {exc}") from exc

    def _extract_xls(self, content: bytes) -> str:
        try:
            import xlrd
        except ModuleNotFoundError as exc:
            raise DocumentExtractionError("xlrd is required for XLS extraction") from exc

        try:
            wb = xlrd.open_workbook(file_contents=content)
            parts: list[str] = []
            for sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
                rows = [
                    [
                        self._format_xlrd_cell(ws.cell(row_idx, col_idx), wb.datemode)
                        for col_idx in range(ws.ncols)
                    ]
                    for row_idx in range(ws.nrows)
                ]
                sheet_text = self._rows_to_structured_text(rows, sheet_name=sheet_name)
                if sheet_text:
                    parts.append(sheet_text)
            return "\n".join(parts)
        except DocumentExtractionError:
            raise
        except Exception as exc:
            raise DocumentExtractionError(f"Failed to extract XLS: {exc}") from exc

    def _format_cell_value(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            if value.hour == 0 and value.minute == 0 and value.second == 0:
                return value.strftime("%Y-%m-%d")
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    def _format_xlrd_cell(self, cell: Any, datemode: int) -> str:
        import xlrd

        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate_as_datetime(cell.value, datemode)
                if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
                    return str(dt.strftime("%Y-%m-%d"))
                return str(dt.strftime("%Y-%m-%d %H:%M:%S"))
            except Exception:
                return str(cell.value)
        return str(cell.value).strip()

    def _rows_to_structured_text(self, rows: list[list[str]], sheet_name: str | None) -> str:
        non_empty = [row for row in rows if any(v.strip() for v in row if v is not None)]
        if len(non_empty) < 2:
            return ""

        headers = non_empty[0]
        data_rows = non_empty[1:]

        lines: list[str] = []
        if sheet_name is not None:
            lines.append(f"[SHEET:{sheet_name}]")

        escaped_headers = [self._escape(h) for h in headers]
        lines.append(f"[HEADERS]:{self._join(escaped_headers)}")

        for idx, row in enumerate(data_rows, start=1):
            padded = (list(row) + [""] * len(headers))[: len(headers)]
            lines.append(f"[ROW:{idx}]:{self._join([self._escape(v) for v in padded])}")

        return "\n".join(lines)

    def _escape(self, value: str) -> str:
        return value.replace("|", "&#124;")

    def _join(self, values: list[str]) -> str:
        return "|".join(values)
