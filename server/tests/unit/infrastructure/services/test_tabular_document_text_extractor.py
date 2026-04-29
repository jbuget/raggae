import io
import sys
import types
from datetime import date, datetime

import pytest

from raggae.domain.exceptions.document_exceptions import DocumentExtractionError
from raggae.infrastructure.services.tabular_document_text_extractor import (
    TabularDocumentTextExtractor,
)


class TestTabularDocumentTextExtractorCsv:
    @pytest.fixture
    def extractor(self) -> TabularDocumentTextExtractor:
        return TabularDocumentTextExtractor()

    async def test_extract_csv_utf8_comma_delimiter(self, extractor: TabularDocumentTextExtractor) -> None:
        # Given
        content = "Produit,Région,Quantité\nWidget A,Nord,1200\nGadget B,Sud,500\n".encode()

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then
        assert "[HEADERS]:Produit|Région|Quantité" in result
        assert "[ROW:1]:Widget A|Nord|1200" in result
        assert "[ROW:2]:Gadget B|Sud|500" in result

    async def test_extract_csv_latin1_fallback(self, extractor: TabularDocumentTextExtractor) -> None:
        # Given
        content = "Produit,Prix\nCafé,2.50\n".encode("latin-1")

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then
        assert "[HEADERS]:Produit|Prix" in result
        assert "Café" in result

    async def test_extract_csv_semicolon_delimiter(self, extractor: TabularDocumentTextExtractor) -> None:
        # Given
        content = b"Col1;Col2;Col3\nA;B;C\nD;E;F\n"

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then
        assert "[HEADERS]:Col1|Col2|Col3" in result
        assert "[ROW:1]:A|B|C" in result

    async def test_extract_csv_tab_delimiter(self, extractor: TabularDocumentTextExtractor) -> None:
        # Given
        content = b"Name\tAge\tCity\nAlice\t30\tParis\n"

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then
        assert "[HEADERS]:Name|Age|City" in result
        assert "[ROW:1]:Alice|30|Paris" in result

    async def test_extract_csv_ignores_empty_rows(self, extractor: TabularDocumentTextExtractor) -> None:
        # Given
        content = b"A,B\nval1,val2\n\n\nval3,val4\n"

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then
        assert "[ROW:1]:val1|val2" in result
        assert "[ROW:2]:val3|val4" in result
        assert result.count("[ROW:") == 2

    async def test_extract_csv_no_sheet_marker(self, extractor: TabularDocumentTextExtractor) -> None:
        # Given
        content = b"H1,H2\nv1,v2\n"

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then
        assert "[SHEET:" not in result

    async def test_extract_csv_escapes_pipe_in_cells(self, extractor: TabularDocumentTextExtractor) -> None:
        # Given
        content = b"Col\nA|B\n"

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then
        assert "&#124;" in result

    async def test_extract_csv_only_headers_returns_empty(
        self, extractor: TabularDocumentTextExtractor
    ) -> None:
        # Given — only one row (headers), no data
        content = b"H1,H2,H3\n"

        # When
        result = await extractor.extract_text("data.csv", content, "text/csv")

        # Then — no structured output (no data rows)
        assert result == ""


class TestTabularDocumentTextExtractorXlsx:
    @pytest.fixture
    def extractor(self) -> TabularDocumentTextExtractor:
        return TabularDocumentTextExtractor()

    @pytest.fixture
    def single_sheet_xlsx(self) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventes"  # type: ignore[union-attr]
        ws.append(["Produit", "Prix"])  # type: ignore[union-attr]
        ws.append(["Widget A", 48.0])  # type: ignore[union-attr]
        ws.append(["Gadget B", 12.5])  # type: ignore[union-attr]
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @pytest.fixture
    def multi_sheet_xlsx(self) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"  # type: ignore[union-attr]
        ws1.append(["Name", "Value"])  # type: ignore[union-attr]
        ws1.append(["Alpha", 1])  # type: ignore[union-attr]
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Key", "Data"])
        ws2.append(["X", "Y"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @pytest.fixture
    def date_xlsx(self) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Label", "Date"])  # type: ignore[union-attr]
        ws.append(["Event", date(2024, 3, 15)])  # type: ignore[union-attr]
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def test_extract_xlsx_single_sheet(
        self,
        extractor: TabularDocumentTextExtractor,
        single_sheet_xlsx: bytes,
    ) -> None:
        # When
        result = await extractor.extract_text(
            "report.xlsx",
            single_sheet_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Then
        assert "[SHEET:Ventes]" in result
        assert "[HEADERS]:Produit|Prix" in result
        assert "[ROW:1]:Widget A|" in result

    async def test_extract_xlsx_multi_sheets_both_present(
        self,
        extractor: TabularDocumentTextExtractor,
        multi_sheet_xlsx: bytes,
    ) -> None:
        # When
        result = await extractor.extract_text(
            "workbook.xlsx",
            multi_sheet_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Then
        assert "[SHEET:Sheet1]" in result
        assert "[SHEET:Sheet2]" in result
        assert "[ROW:1]:Alpha|1" in result
        assert "[ROW:1]:X|Y" in result

    async def test_extract_xlsx_date_formatted_as_yyyy_mm_dd(
        self,
        extractor: TabularDocumentTextExtractor,
        date_xlsx: bytes,
    ) -> None:
        # When
        result = await extractor.extract_text(
            "dates.xlsx", date_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Then
        assert "2024-03-15" in result


_XL_CELL_EMPTY = 0
_XL_CELL_TEXT = 1
_XL_CELL_NUMBER = 2
_XL_CELL_DATE = 3


class TestTabularDocumentTextExtractorXls:
    @pytest.fixture
    def extractor(self) -> TabularDocumentTextExtractor:
        return TabularDocumentTextExtractor()

    def _make_fake_xlrd(
        self,
        sheet_name: str = "Feuille1",
        rows: list[list[tuple[int, object]]] | None = None,
    ) -> types.ModuleType:
        if rows is None:
            rows = [
                [(_XL_CELL_TEXT, "Nom"), (_XL_CELL_TEXT, "Valeur")],
                [(_XL_CELL_TEXT, "Alpha"), (_XL_CELL_NUMBER, "42")],
            ]

        class FakeCell:
            def __init__(self, ctype: int, value: object) -> None:
                self.ctype = ctype
                self.value = value

        sheet_data = [[FakeCell(ctype, val) for ctype, val in row] for row in rows]

        class FakeSheet:
            def __init__(self) -> None:
                self.nrows = len(sheet_data)
                self.ncols = len(sheet_data[0]) if sheet_data else 0

            def cell(self, row: int, col: int) -> FakeCell:
                return sheet_data[row][col]  # type: ignore[return-value]

        _sheet_name = sheet_name
        _sheet = FakeSheet()

        class FakeWorkbook:
            datemode = 0

            def sheet_names(self) -> list[str]:
                return [_sheet_name]

            def sheet_by_name(self, name: str) -> FakeSheet:
                return _sheet

        fake = types.ModuleType("xlrd")
        fake.XL_CELL_EMPTY = _XL_CELL_EMPTY  # type: ignore[attr-defined]
        fake.XL_CELL_DATE = _XL_CELL_DATE  # type: ignore[attr-defined]
        fake.open_workbook = lambda file_contents: FakeWorkbook()  # type: ignore[attr-defined]
        fake.xldate_as_datetime = lambda v, dm: datetime(2024, 1, 15, 0, 0, 0)  # type: ignore[attr-defined]
        return fake

    async def test_extract_xls_with_mocked_xlrd(
        self, extractor: TabularDocumentTextExtractor, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        # Given
        fake_xlrd = self._make_fake_xlrd()
        monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)

        # When
        result = await extractor.extract_text("legacy.xls", b"fake", "application/vnd.ms-excel")

        # Then
        assert "[SHEET:Feuille1]" in result
        assert "[HEADERS]:Nom|Valeur" in result
        assert "[ROW:1]:Alpha|42" in result

    async def test_extract_xls_date_formatted_as_yyyy_mm_dd(
        self, extractor: TabularDocumentTextExtractor, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        # Given
        fake_xlrd = self._make_fake_xlrd(
            sheet_name="Data",
            rows=[
                [(_XL_CELL_TEXT, "Label"), (_XL_CELL_TEXT, "Date")],
                [(_XL_CELL_TEXT, "Event"), (_XL_CELL_DATE, 45306.0)],
            ],
        )
        monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)

        # When
        result = await extractor.extract_text("legacy.xls", b"fake", "application/vnd.ms-excel")

        # Then
        assert "2024-01-15" in result

    async def test_extract_unsupported_extension_raises(
        self, extractor: TabularDocumentTextExtractor
    ) -> None:
        # When / Then
        with pytest.raises(DocumentExtractionError):
            await extractor.extract_text(
                "file.ods", b"data", "application/vnd.oasis.opendocument.spreadsheet"
            )
